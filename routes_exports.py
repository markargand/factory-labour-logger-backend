# routes_exports.py
# Export labour data to XLSX and PDF.
#
#  - GET /export/xlsx?week=2025-W47  -> Excel with 3 sheets:
#       * Entries
#       * By employee
#       * By project
#  - GET /export/pdf?week=2025-W47   -> Simple PDF summary

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session
from datetime import date as date_cls
from io import BytesIO
from collections import defaultdict

from database import get_db  # uses your existing database.py

router = APIRouter(prefix="/export", tags=["export"])


def iso_week_range(week_id: str):
    """
    Convert '2025-W47' into start/end date strings (YYYY-MM-DD).
    If invalid, returns (None, None).
    """
    try:
        year_str, week_str = week_id.split("-W")
        year = int(year_str)
        week = int(week_str)
        start = date_cls.fromisocalendar(year, week, 1)  # Monday
        end = date_cls.fromisocalendar(year, week, 7)    # Sunday
        return start.isoformat(), end.isoformat()
    except Exception:
        return None, None


def fetch_entries(db: Session, week: str | None):
    """
    Fetch entries joined with employees & projects.
    If week is provided (e.g., '2025-W47'), filter by that ISO week.
    """
    params = {}
    where = ""
    if week:
        start_date, end_date = iso_week_range(week)
        if start_date and end_date:
            where = "WHERE e.date BETWEEN :start AND :end"
            params["start"] = start_date
            params["end"] = end_date

    rows = db.execute(text(f"""
        SELECT
          e.*,
          emp.name  AS employee_name,
          proj.code AS project_code,
          proj.name AS project_name
        FROM entries e
        LEFT JOIN employees emp ON e.employee_id = emp.id
        LEFT JOIN projects proj ON e.project_id  = proj.id
        {where}
        ORDER BY e.date, employee_name
    """), params).mappings().all()
    return rows


@router.get("/xlsx")
def export_xlsx(
    week: str | None = Query(None, description="ISO week like 2025-W47"),
    db: Session = Depends(get_db),
):
    """
    Generate an Excel workbook with:
      - Entries     (one row per entry)
      - By employee (total hours)
      - By project  (total hours)
    """
    from openpyxl import Workbook

    rows = fetch_entries(db, week)

    wb = Workbook()
    ws_entries = wb.active
    ws_entries.title = "Entries"

    # Sheet 1: Entries
    headers = [
        "Date", "Employee", "Project Code", "Project Name",
        "Work Type", "Start", "End", "Break (min)",
        "Hours", "Status", "Locked"
    ]
    ws_entries.append(headers)

    for r in rows:
        ws_entries.append([
            r["date"],
            r.get("employee_name"),
            r.get("project_code"),
            r.get("project_name"),
            r.get("work_type"),
            r.get("start"),
            r.get("end"),
            r.get("break_min"),
            r.get("hours"),
            r.get("status"),
            "Yes" if r.get("locked") else "No",
        ])

    # Sheet 2: By employee
    emp_totals = defaultdict(float)
    for r in rows:
        hrs = r.get("hours") or 0
        name = r.get("employee_name") or r.get("employee_id")
        emp_totals[name] += hrs

    ws_emp = wb.create_sheet(title="By employee")
    ws_emp.append(["Employee", "Total hours"])
    for name, hours in sorted(emp_totals.items()):
        ws_emp.append([name, float(hours)])

    # Sheet 3: By project
    proj_totals = defaultdict(float)
    for r in rows:
        hrs = r.get("hours") or 0
        code = r.get("project_code") or ""
        name = r.get("project_name") or ""
        key = f"{code} {name}".strip()
        proj_totals[key] += hrs

    ws_proj = wb.create_sheet(title="By project")
    ws_proj.append(["Project", "Total hours"])
    for key, hours in sorted(proj_totals.items()):
        ws_proj.append([key, float(hours)])

    # Write workbook to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename_week = week or "all"
    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="labour-report-{filename_week}.xlsx"'
        },
    )


@router.get("/pdf")
def export_pdf(
    week: str | None = Query(None, description="ISO week like 2025-W47"),
    db: Session = Depends(get_db),
):
    """
    Generate a simple PDF summary:
      - Totals by employee
      - Totals by project
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    rows = fetch_entries(db, week)

    emp_totals = defaultdict(float)
    proj_totals = defaultdict(float)

    for r in rows:
        hrs = r.get("hours") or 0
        # Employee key
        ename = r.get("employee_name") or r.get("employee_id")
        emp_totals[ename] += hrs

        # Project key
        code = r.get("project_code") or ""
        pname = r.get("project_name") or ""
        pkey = f"{code} {pname}".strip()
        proj_totals[pkey] += hrs

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    title = "Factory Labour Summary"
    c.setFont("Helvetica-Bold", 16)
    c.drawString(40, height - 40, title)

    if week:
        c.setFont("Helvetica", 10)
        c.drawString(40, height - 60, f"Week: {week}")

    y = height - 90
    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "By employee")
    y -= 20
    c.setFont("Helvetica", 10)
    for name, hours in sorted(emp_totals.items()):
        if y < 60:
            c.showPage()
            y = height - 40
        c.drawString(50, y, f"{name}: {hours:.2f} h")
        y -= 15

    y -= 15
    if y < 60:
        c.showPage()
        y = height - 40

    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "By project")
    y -= 20
    c.setFont("Helvetica", 10)
    for key, hours in sorted(proj_totals.items()):
        if y < 60:
            c.showPage()
            y = height - 40
        c.drawString(50, y, f"{key}: {hours:.2f} h")
        y -= 15

    c.showPage()
    c.save()
    buffer.seek(0)

    filename_week = week or "all"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="labour-summary-{filename_week}.pdf"'
        },
    )
